#!/usr/bin/env python3
"""Lädt EEX-TTF-NDI- und EUA-Auktionsdaten, vereinheitlicht und archiviert sie.

Wichtiger Lizenzhinweis:
Die EEX gestattet eine systematische öffentliche Weiterverbreitung erheblicher
Datenmengen nur mit ausdrücklicher Genehmigung. Deshalb schreibt dieses Skript
EEX-Daten nur bei privaten Repositories dauerhaft in ``Data``. Bei öffentlichen
Repositories landen die Daten ausschließlich im nicht versionierten Ordner
``PrivateData`` und im Workflow-Artefakt.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import ssl
import sys
import time
import urllib.error
import urllib.request
import zipfile
from collections.abc import Iterable
from datetime import date, datetime, time as dt_time, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PUBLIC_DATA_DIR = ROOT / "Data"
PRIVATE_DATA_DIR = ROOT / "PrivateData"
DOWNLOAD_DIR = ROOT / "download"
ARTIFACT_DIR = ROOT / "_artifacts"
TIMEZONE = ZoneInfo("Europe/Berlin")

REPOSITORY_PRIVATE = os.getenv("REPOSITORY_PRIVATE", "false").lower() in {
    "1", "true", "yes", "ja"
}
FORCE_PUBLIC_COMMIT = os.getenv("FORCE_EEX_PUBLIC_COMMIT", "false").lower() in {
    "1", "true", "yes", "ja"
}
FULL_REFRESH = os.getenv("FULL_REFRESH", "false").lower() in {"1", "true", "yes", "ja"}

OUTPUT_DIR = PUBLIC_DATA_DIR if (REPOSITORY_PRIVATE or FORCE_PUBLIC_COMMIT) else PRIVATE_DATA_DIR
ZIP_DIR = DOWNLOAD_DIR if (REPOSITORY_PRIVATE or FORCE_PUBLIC_COMMIT) else ARTIFACT_DIR

TTF_NDI_URL = "https://gasandregistry.eex.com/Gas/NDI/NDI_45_Days.csv"
TTF_ISSUER_URL = "https://secure.globalsign.com/cacert/gsgccr3dvtlsca2020.crt"
CURRENT_YEAR = datetime.now(TIMEZONE).year
EUA_CURRENT_URL = (
    "https://public.eex-group.com/eex/eua-auction-report/"
    f"emission-spot-primary-market-auction-report-{CURRENT_YEAR}-data.xlsx"
)
EUA_ARCHIVE_URLS = (
    "https://www.eex.com/fileadmin/EEX/Downloads/Markets/Environmentals/"
    "EUA_Emission_Spot_Primary_Market_Auction_Report/Archive_Reports/"
    f"emission-spot-primary-market-auction-report-2012-{CURRENT_YEAR - 1}-data.zip",
    "https://www.eex.com/fileadmin/EEX/Downloads/Markets/Environmentals/"
    "EUA_Emission_Spot_Primary_Market_Auction_Report/Archive_Reports/"
    "emission-spot-primary-market-auction-report-2012-2025-data.zip",
)

USER_AGENT = "Mozilla/5.0 SMARD-EEX-GitHub-Actions-Downloader/1.0"
DOWNLOAD_TIMEOUT = 180
_TTF_SSL_CONTEXT: ssl.SSLContext | None = None




def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)

def utc_now_text() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def ttf_ssl_context() -> ssl.SSLContext:
    """Ergänzt die vom EEX-Server nicht mitgesendete GlobalSign-Zwischen-CA.

    Die TLS-Prüfung bleibt vollständig aktiv. Das Zwischenzertifikat wird
    seinerseits per HTTPS von GlobalSign geladen und nur als zusätzliche
    Ketteninformation verwendet; die Vertrauensanker bleiben die System-CAs.
    """
    global _TTF_SSL_CONTEXT
    if _TTF_SSL_CONTEXT is not None:
        return _TTF_SSL_CONTEXT

    with urllib.request.urlopen(TTF_ISSUER_URL, timeout=DOWNLOAD_TIMEOUT) as response:
        issuer_der = response.read()
    if len(issuer_der) < 500:
        raise ValueError("GlobalSign-Zwischenzertifikat ist ungewöhnlich klein.")

    context = ssl.create_default_context()
    context.load_verify_locations(cadata=ssl.DER_cert_to_PEM_cert(issuer_der))
    _TTF_SSL_CONTEXT = context
    return context


def download_bytes(url: str, label: str, attempts: int = 4) -> bytes:
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": USER_AGENT,
                "Accept": "*/*",
                "Cache-Control": "no-cache",
            },
        )
        try:
            context = ttf_ssl_context() if url == TTF_NDI_URL else None
            with urllib.request.urlopen(
                req, timeout=DOWNLOAD_TIMEOUT, context=context
            ) as response:
                payload = response.read()
            if len(payload) < 100:
                raise ValueError(f"Antwort ist ungewöhnlich klein ({len(payload)} Byte).")
            return payload
        except (urllib.error.URLError, TimeoutError, ValueError) as exc:
            last_error = exc
            if attempt < attempts:
                wait = attempt * 8
                print(f"  {label}: Versuch {attempt} fehlgeschlagen: {exc}; neuer Versuch in {wait}s")
                time.sleep(wait)
    raise RuntimeError(f"{label} konnte nicht geladen werden: {last_error}")


def download_first(urls: Iterable[str], label: str) -> bytes:
    errors: list[str] = []
    for url in dict.fromkeys(urls):
        try:
            return download_bytes(url, label)
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    raise RuntimeError(f"{label}: keine Downloadadresse funktionierte. " + " | ".join(errors))


def decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Datei konnte nicht als Text dekodiert werden.")


def norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = (
        text.replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
        .replace("₂", "2")
    )
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if number == number else None
    text = str(value).strip().replace("€", "").replace("EUR", "").replace("eur", "")
    text = text.replace("\u00a0", "").replace(" ", "")
    if not text or text.lower() in {"-", "n/a", "na", "nan", "null"}:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9+\-.eE]", "", text)
    try:
        number = float(text)
    except ValueError:
        return None
    return number if number == number else None


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"\s+00:00:00$", "", text)
    for fmt in (
        "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y",
        "%Y/%m/%d", "%d-%m-%Y", "%d %b %Y", "%d %B %Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.search(r"(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    match = re.search(r"(\d{1,2})[-/.](\d{1,2})[-/.](20\d{2})", text)
    if match:
        try:
            return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except ValueError:
            return None
    return None


def sniff_delimiter(text: str) -> str:
    lines = [line for line in text.splitlines() if line.strip()][:10]
    if not lines:
        return ";"
    # Die Kopfzeile ist zuverlässiger als das Gesamtsample, weil Dezimalkommas
    # sonst fälschlich als CSV-Trennzeichen erkannt werden können.
    first = lines[0]
    first_counts = {d: first.count(d) for d in (";", "\t", "|", ",")}
    if max(first_counts.values()) > 0:
        winner = max(first_counts, key=first_counts.get)
        if first_counts[winner] >= 1:
            return winner
    sample = "\n".join(lines)
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        counts = {d: sample.count(d) for d in (";", "\t", "|", ",")}
        return max(counts, key=counts.get)


def combined_headers(rows: list[list[Any]], header_index: int) -> list[str]:
    width = max(len(rows[i]) for i in range(max(0, header_index - 2), header_index + 1))
    result: list[str] = []
    for col in range(width):
        parts: list[str] = []
        for row_index in range(max(0, header_index - 2), header_index + 1):
            cell = rows[row_index][col] if col < len(rows[row_index]) else ""
            n = norm(cell)
            if n and n not in parts:
                parts.append(n)
        result.append(" ".join(parts))
    return result


def choose_header(rows: list[list[Any]], product_token: str) -> tuple[int, list[str]]:
    best: tuple[int, int, list[str]] | None = None
    for index in range(min(40, len(rows))):
        current = [norm(cell) for cell in rows[index]]
        current_joined = " | ".join(current)
        headers = combined_headers(rows, index)
        score = 0
        if "ttf" in current_joined:
            score += 6
        if product_token in current_joined:
            score += 5
        if any(token in current_joined for token in ("date", "delivery", "gas day", "gasday", "liefer")):
            score += 4
        if any(token in current_joined for token in ("price", "index", "value", "preis")):
            score += 3
        # Echte Datenzeilen enthalten häufig bereits TTF, Datum und Preis. Sie dürfen
        # nicht als Tabellenkopf ausgewählt werden.
        if any(parse_date(cell) is not None for cell in rows[index]):
            score -= 10
        if score > 0 and (best is None or score > best[0]):
            best = (score, index, headers)
    if best is None:
        raise ValueError("Kein plausibler Tabellenkopf gefunden.")
    return best[1], best[2]


def find_column(headers: list[str], tokens: Iterable[str], exclude: Iterable[str] = ()) -> int | None:
    token_list = tuple(tokens)
    exclude_list = tuple(exclude)
    candidates: list[tuple[int, int]] = []
    for idx, header in enumerate(headers):
        if any(ex in header for ex in exclude_list):
            continue
        hits = sum(1 for token in token_list if token in header)
        if hits:
            candidates.append((hits, idx))
    return max(candidates)[1] if candidates else None


def parse_ttf_ndi(raw: bytes) -> dict[date, float]:
    text = decode_text(raw).replace("\r\n", "\n").replace("\r", "\n")
    delimiter = sniff_delimiter(text)
    rows = [row for row in csv.reader(io.StringIO(text), delimiter=delimiter) if any(str(c).strip() for c in row)]
    if len(rows) < 3:
        raise ValueError("TTF-NDI-Datei enthält zu wenige Zeilen.")
    header_index, headers = choose_header(rows, "ndi")
    date_idx = find_column(headers, ("delivery date", "gas day", "gasday", "date", "lieferdatum", "delivery"))
    if date_idx is None:
        raise ValueError(f"Datumsfeld nicht gefunden. Erkannte Spalten: {headers}")

    # Breites Format: eine eigene TTF-Spalte.
    ttf_candidates: list[tuple[int, int]] = []
    for idx, header in enumerate(headers):
        if "ttf" in header:
            score = 1 + (4 if "ndi" in header else 0) + (2 if "price" in header or "preis" in header else 0)
            ttf_candidates.append((score, idx))
    ttf_idx = max(ttf_candidates)[1] if ttf_candidates else None

    # Langes Format: Hub/Market Area und Preis stehen in getrennten Spalten.
    hub_idx = find_column(headers, ("market area", "marketarea", "hub", "zone", "index name", "product"))
    price_idx = find_column(headers, ("price", "index value", "value", "preis", "ndi"), exclude=("date",))

    parsed: dict[date, float] = {}
    for row in rows[header_index + 1 :]:
        if date_idx >= len(row):
            continue
        day = parse_date(row[date_idx])
        if day is None:
            continue
        price: float | None = None
        if ttf_idx is not None and ttf_idx < len(row):
            price = parse_number(row[ttf_idx])
        if price is None and hub_idx is not None and price_idx is not None:
            if hub_idx < len(row) and "ttf" in norm(row[hub_idx]) and price_idx < len(row):
                price = parse_number(row[price_idx])
        if price is None and "ttf" in " ".join(norm(c) for c in row):
            numbers = [parse_number(c) for c in row]
            plausible = [n for n in numbers if n is not None and -500 <= n <= 1000]
            if plausible:
                price = plausible[-1]
        if price is not None and -500 <= price <= 1000:
            parsed[day] = price
    if len(parsed) < 5:
        raise ValueError(f"Nur {len(parsed)} TTF-NDI-Werte erkannt; CSV-Format möglicherweise geändert.")
    return parsed


def find_xlsx_header(rows: list[list[Any]]) -> tuple[int, int, int] | None:
    best: tuple[int, int, int, int] | None = None
    for i in range(min(100, len(rows))):
        headers = combined_headers(rows, i)
        # Nicht nur nach "auction" suchen: Sonst gewinnt in den aktuellen
        # EEX-Dateien fälschlich die Spalte "Auction Name" statt "Date".
        date_candidates: list[tuple[int, int]] = []
        price_candidates: list[tuple[int, int]] = []
        for idx, header in enumerate(headers):
            if header == "date" or header.endswith(" date") or header.startswith("date "):
                date_candidates.append((10, idx))
            elif "auction date" in header or "auction day" in header or "auktionsdatum" in header:
                date_candidates.append((9, idx))

            if any(token in header for token in (
                "auction price", "clearing price", "settlement price",
                "auction clearing", "auktionspreis",
            )) and not any(token in header for token in (
                "minimum", "maximum", "mean", "median", "volume", "bid",
            )):
                price_candidates.append((10, idx))
            elif "price eur" in header or header == "preis":
                price_candidates.append((5, idx))

        date_idx = max(date_candidates)[1] if date_candidates else None
        price_idx = max(price_candidates)[1] if price_candidates else None
        if date_idx is None or price_idx is None or date_idx == price_idx:
            continue
        score = 0
        joined = " | ".join(headers)
        if "auction price" in joined or "clearing price" in joined:
            score += 6
        if "auction date" in joined:
            score += 5
        if "eua" in joined or "emission" in joined:
            score += 2
        if best is None or score > best[0]:
            best = (score, i, date_idx, price_idx)
    return None if best is None else (best[1], best[2], best[3])


def parse_eua_workbook(raw: bytes, source_name: str) -> dict[date, float]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    parsed: dict[date, float] = {}
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        header = find_xlsx_header(rows)
        if header is None:
            continue
        header_index, date_idx, price_idx = header
        blanks = 0
        for row in rows[header_index + 1 :]:
            if not any(value not in (None, "") for value in row):
                blanks += 1
                if blanks >= 20:
                    break
                continue
            blanks = 0
            if date_idx >= len(row) or price_idx >= len(row):
                continue
            day = parse_date(row[date_idx])
            price = parse_number(row[price_idx])
            if day is None or price is None:
                continue
            if date(2010, 1, 1) <= day <= date.today() + timedelta(days=10) and 1 <= price <= 500:
                parsed[day] = price
    if not parsed:
        raise ValueError(f"In {source_name} wurden keine EUA-Auktionspreise erkannt.")
    return parsed


def parse_eua_archive(raw: bytes) -> dict[date, float]:
    parsed: dict[date, float] = {}
    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        names = [name for name in archive.namelist() if name.lower().endswith((".xlsx", ".xlsm"))]
        if not names:
            raise ValueError("EUA-Archiv enthält keine XLSX-Dateien.")
        for name in names:
            try:
                parsed.update(parse_eua_workbook(archive.read(name), name))
            except Exception as exc:  # einzelne Altdatei darf den Gesamtabruf nicht blockieren
                print(f"  Warnung: {name} konnte nicht ausgewertet werden: {exc}")
    if not parsed:
        raise ValueError("Im EUA-Archiv wurden keine Preise erkannt.")
    return parsed


def read_standard_csv(path: Path, date_column: str, price_column: str) -> dict[date, float]:
    if not path.exists() or path.stat().st_size < 30:
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        result: dict[date, float] = {}
        for row in reader:
            day = parse_date(row.get(date_column))
            price = parse_number(row.get(price_column))
            if day is not None and price is not None:
                result[day] = price
        return result


def ndi_available_from(delivery_day: date) -> str:
    publication_day = delivery_day - timedelta(days=1)
    local = datetime.combine(publication_day, dt_time(18, 0), tzinfo=TIMEZONE)
    return local.isoformat()


def write_ttf(path: Path, values: dict[date, float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fetched = utc_now_text()
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";", lineterminator="\n")
        writer.writerow([
            "Lieferdatum", "TTF_NDI_EUR_MWh", "Verfuegbar_ab_Europe_Berlin",
            "Produkt", "Quelle", "Letzter_Abruf_UTC",
        ])
        for day, price in sorted(values.items()):
            writer.writerow([
                day.isoformat(), f"{price:.6f}".replace(".", ","), ndi_available_from(day),
                "EEX TTF Next Day Index (NDI)", "EEX", fetched,
            ])


def write_eua_raw(path: Path, values: dict[date, float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fetched = utc_now_text()
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";", lineterminator="\n")
        writer.writerow(["Auktionsdatum", "EUA_Auktionspreis_EUR_tCO2", "Produkt", "Quelle", "Letzter_Abruf_UTC"])
        for day, price in sorted(values.items()):
            writer.writerow([
                day.isoformat(), f"{price:.6f}".replace(".", ","),
                "EUA Primary Market Auction", "EEX", fetched,
            ])


def write_eua_daily(path: Path, values: dict[date, float], start: date = date(2022, 1, 1)) -> None:
    filtered = {d: p for d, p in values.items() if d <= date.today()}
    if not filtered:
        raise ValueError("Keine EUA-Werte für Tagesreihe vorhanden.")
    sorted_dates = sorted(filtered)
    first = max(start, sorted_dates[0])
    last = date.today()
    current_price: float | None = None
    current_source_date: date | None = None
    idx = 0
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";", lineterminator="\n")
        writer.writerow([
            "Datum", "EUA_Auktionspreis_zuletzt_bekannt_EUR_tCO2",
            "Zugehoeriges_Auktionsdatum", "Methode", "Quelle",
        ])
        day = first
        while day <= last:
            while idx < len(sorted_dates) and sorted_dates[idx] <= day:
                current_source_date = sorted_dates[idx]
                current_price = filtered[current_source_date]
                idx += 1
            if current_price is not None and current_source_date is not None:
                writer.writerow([
                    day.isoformat(), f"{current_price:.6f}".replace(".", ","),
                    current_source_date.isoformat(), "vorwaerts fortgeschrieben", "EEX",
                ])
            day += timedelta(days=1)


def build_zip(files: list[Path]) -> Path:
    ZIP_DIR.mkdir(parents=True, exist_ok=True)
    zip_path = ZIP_DIR / "EEX_Gas_CO2_aktuell.zip"
    notice = ZIP_DIR / "EEX_QUELLE_UND_LIZENZHINWEIS.txt"
    notice.write_text(
        "Quellen:\n"
        "- European Energy Exchange AG (EEX): TTF Next Day Index\n"
        "- European Energy Exchange AG (EEX): EUA Primary Market Auction Report\n\n"
        "Wichtiger Lizenzhinweis:\n"
        "Die systematische öffentliche Weiterverbreitung erheblicher EEX-Datenmengen "
        "ist nur mit ausdrücklicher Genehmigung der EEX gestattet. Dieses Paket ist "
        "für interne Analysezwecke vorgesehen.\n\n"
        "Die EUA-Tagesreihe ist aus den einzelnen Auktionspreisen vorwärts fortgeschrieben. "
        "Für zeitlich saubere Prognosen muss zusätzlich der tatsächliche Veröffentlichungszeitpunkt "
        "der jeweiligen Information berücksichtigt werden.\n",
        encoding="utf-8",
    )
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for file in files:
            archive.write(file, arcname=file.name)
        archive.write(notice, arcname=notice.name)
    return zip_path


def update_ttf() -> Path:
    print("\nTTF Next Day Index wird geladen …")
    path = OUTPUT_DIR / "TTF_NDI_aktuell.csv"
    existing = read_standard_csv(path, "Lieferdatum", "TTF_NDI_EUR_MWh")
    raw = download_bytes(TTF_NDI_URL, "TTF NDI")
    current = parse_ttf_ndi(raw)
    existing.update(current)
    write_ttf(path, existing)
    print(f"  Gespeichert: {display_path(path)} ({len(existing)} Liefertage, bis {max(existing)})")
    return path


def update_eua() -> tuple[Path, Path]:
    print("\nEUA-Auktionspreise werden geladen …")
    raw_path = OUTPUT_DIR / "EUA_Auktionspreis_aktuell.csv"
    daily_path = OUTPUT_DIR / "EUA_Auktionspreis_taeglich.csv"
    existing = read_standard_csv(raw_path, "Auktionsdatum", "EUA_Auktionspreis_EUR_tCO2")

    # Historie wird beim ersten Lauf bzw. Full Refresh neu eingelesen.
    if FULL_REFRESH or not existing or min(existing, default=date.max) > date(2022, 1, 15):
        archive = download_first(EUA_ARCHIVE_URLS, f"EUA-Archiv bis {CURRENT_YEAR - 1}")
        existing.update(parse_eua_archive(archive))
        print(f"  Historisches Archiv: {len(existing)} Auktionswerte erkannt")

    current = download_bytes(EUA_CURRENT_URL, f"EUA-Auktionsreport {CURRENT_YEAR}")
    existing.update(parse_eua_workbook(current, f"EUA-Auktionsreport {CURRENT_YEAR}"))
    existing = {d: p for d, p in existing.items() if d >= date(2022, 1, 1)}
    write_eua_raw(raw_path, existing)
    write_eua_daily(daily_path, existing)
    print(f"  Gespeichert: {display_path(raw_path)} ({len(existing)} Auktionen, bis {max(existing)})")
    print(f"  Gespeichert: {display_path(daily_path)}")
    return raw_path, daily_path


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"EEX-Ausgabemodus: {'dauerhaft im privaten Repository' if OUTPUT_DIR == PUBLIC_DATA_DIR else 'nur Workflow-Artefakt (Repository ist öffentlich)'}")
    if FORCE_PUBLIC_COMMIT and not REPOSITORY_PRIVATE:
        print("WARNUNG: FORCE_EEX_PUBLIC_COMMIT ist aktiv. EEX-Lizenzbedingungen eigenverantwortlich prüfen.")

    produced: list[Path] = []
    errors: list[str] = []
    for label, func in (("TTF", update_ttf), ("EUA", update_eua)):
        try:
            result = func()
            produced.extend(result if isinstance(result, tuple) else (result,))
        except Exception as exc:
            message = f"{label}-Abruf fehlgeschlagen: {exc}"
            print("FEHLER:", message)
            errors.append(message)

    if produced:
        zip_path = build_zip(produced)
        print(f"\nEEX-ZIP erstellt: {display_path(zip_path)}")
    else:
        print("Keine EEX-Datei konnte erzeugt werden.")

    status = {
        "timestamp_utc": utc_now_text(),
        "repository_private": REPOSITORY_PRIVATE,
        "output_directory": display_path(OUTPUT_DIR),
        "produced_files": [display_path(p) for p in produced],
        "errors": errors,
    }
    ZIP_DIR.mkdir(parents=True, exist_ok=True)
    (ZIP_DIR / "EEX_Status.json").write_text(json.dumps(status, indent=2, ensure_ascii=False), encoding="utf-8")

    # Bereits vorhandene Daten dürfen bei einem vorübergehenden Quellenausfall erhalten bleiben.
    # Der Workflow wird nur dann hart abgebrochen, wenn überhaupt keine Datei erzeugt werden konnte.
    return 0 if produced else 1


if __name__ == "__main__":
    sys.exit(main())
